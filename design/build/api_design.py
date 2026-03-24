import os, time, requests
from dotenv import load_dotenv
import openpyxl

load_dotenv()
API_KEY = os.getenv("KIPRISPLUS_API_KEY")
assert API_KEY, "KIPRISPLUS_API_KEY 없음 (.env 확인)"

# 출원번호 들어있는 엑셀 파일 로드
excel_file = r"C:\Users\playdata2\Desktop\SKN_AI_20\SKN20-FINAL-2TEAM\data\출원번호\2025_2026.xlsx"  #저장한 경로로 바꿔주세요!
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb.active

# C열(C9부터) 출원번호 추출 
application_numbers = []
for row in range(197, ws.max_row + 1):
    cell_value = ws[f'C{row}'].value
    if cell_value:
        application_numbers.append(str(cell_value).strip())

print(f"📊 {excel_file} 파일에서 {len(application_numbers)}개의 출원번호를 읽었습니다.\n")

# 서지상세정보 API 호출
base_url = "http://plus.kipris.or.kr/kipo-api/kipi/designInfoSearchService/getBibliographyDetailInfoSearch"

# **중요** 연도수로 폴더명 지정하기! 
os.makedirs("../data/xml/2025_2026", exist_ok=True) # xml 저장할 폴더 생성

success_count = 0
fail_count = 0

for idx, app_num in enumerate(application_numbers, 1):
    try:
        params = {
            "applicationNumber": app_num,  # 출원번호
            "ServiceKey": API_KEY,  # API 키
        }
        
        t0 = time.time()
        r = requests.get(base_url, params=params, timeout=20)
        latency_ms = int((time.time() - t0) * 1000)
        
        print(f"[{idx}/{len(application_numbers)}] {app_num} - status: {r.status_code}, latency: {latency_ms}ms")
        
        # XML 파일로 저장
        if r.status_code == 200:
            file_path = f"../data/xml/2025_2026/{app_num}.xml" #폴더명 데이터 맞춰서 변경
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(r.text)
            print(f"         ✅ 저장: {file_path}")
            success_count += 1
        else:
            print(f"         ❌ 오류: {r.status_code}")
            fail_count += 1
            
        time.sleep(0.5)  # API 요청 간 딜레이
        
    except requests.exceptions.RequestException as e:
        print(f"[{idx}/{len(application_numbers)}] {app_num} - Request failed: {type(e).__name__}")
        fail_count += 1

print(f"\n✅ 완료: {success_count}개 저장, {fail_count}개 실패")
print(f"📁 모든 서지상세정보 XML 파일이 '2025_2026' 폴더에 저장되었습니다.") 


